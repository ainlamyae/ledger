"""
Build a scrubbed, demo-data template workbook from the personal production
export, for use as the multi-user Google Sheets template (TEMPLATE_SPREADSHEET_ID).

Reads SOURCE_PATH read-only. Never writes back to it. Writes OUTPUT_PATH only.
"""

import datetime
import openpyxl

SOURCE_PATH = r"C:\Users\AlanKay\Downloads\Ledger Database.xlsx"
OUTPUT_PATH = r"C:\Users\AlanKay\Downloads\Ledger Template.xlsx"

# (date, account, payee, description, amount, category)
# Categories chosen to match every category column in 'Monthly Summary' (D1:L1):
# Housing, Transportation, Grocery, Personal, Household, Fee, Medical,
# Application, Donation -- plus one Income row and one RRSP contribution pair
# so the 'Account Balance' RRSP formulas have something to sum.
DEMO_TRANSACTIONS = [
    (datetime.datetime(2025, 4, 1), "Checking", "Employer Inc",        "Payroll Deposit",            3200.00, "Income"),
    (datetime.datetime(2025, 4, 2), "Checking", "Generic Landlord",    "Home Rent",                  -1200.00, "Housing"),
    (datetime.datetime(2025, 4, 3), "Credit Card", "Generic Grocery Store", "Weekly Groceries",        -85.40, "Grocery"),
    (datetime.datetime(2025, 4, 5), "Credit Card", "City Transit",      "Monthly Transit Pass",        -95.00, "Transportation"),
    (datetime.datetime(2025, 4, 7), "Cash", "Sample Cafe",              "Coffee",                       -4.50, "Personal"),
    (datetime.datetime(2025, 4, 10), "Credit Card", "Generic Pharmacy", "Toiletries",                   -22.10, "Household"),
    (datetime.datetime(2025, 4, 12), "Checking", "Sample Bank",         "Monthly Account Fee",          -5.00, "Fee"),
    (datetime.datetime(2025, 4, 15), "Credit Card", "Generic Clinic",   "Doctor Visit Copay",          -40.00, "Medical"),
    (datetime.datetime(2025, 4, 18), "Checking", "Visa Office",         "Application Fee",            -150.00, "Application"),
    (datetime.datetime(2025, 4, 20), "Checking", "Local Charity",       "Monthly Donation",             -25.00, "Donation"),
    (datetime.datetime(2025, 4, 21), "RRSP (Member)", "Sample Plan Provider", "RRSP Contribution (Member)", 200.00, "Saving"),
    (datetime.datetime(2025, 4, 21), "RRSP (Employer)", "Sample Plan Provider", "RRSP Contribution (Employer)", 100.00, "Saving"),
    (datetime.datetime(2025, 5, 1), "Checking", "Employer Inc",        "Payroll Deposit",            3200.00, "Income"),
    (datetime.datetime(2025, 5, 2), "Checking", "Generic Landlord",    "Home Rent",                  -1200.00, "Housing"),
    (datetime.datetime(2025, 5, 4), "Credit Card", "Generic Grocery Store", "Weekly Groceries",        -91.25, "Grocery"),
    (datetime.datetime(2025, 5, 6), "Credit Card", "City Transit",      "Monthly Transit Pass",        -95.00, "Transportation"),
    (datetime.datetime(2025, 5, 9), "Cash", "Sample Cafe",              "Coffee",                       -4.50, "Personal"),
    (datetime.datetime(2025, 5, 13), "Credit Card", "Generic Pharmacy", "Household Supplies",          -18.75, "Household"),
    (datetime.datetime(2025, 5, 16), "Checking", "Generic Restaurant",  "Dinner Out",                  -38.00, "Personal"),
    (datetime.datetime(2025, 5, 20), "Checking", "Local Charity",       "Monthly Donation",             -25.00, "Donation"),
]

NEW_LAST_ROW = 1 + len(DEMO_TRANSACTIONS)  # header + demo rows


def rebuild_transactions(wb):
    ws = wb["Transactions"]
    ws.delete_rows(2, ws.max_row - 1)
    for i, row in enumerate(DEMO_TRANSACTIONS, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=value)


def rebuild_account_balance(wb):
    ws = wb["Account Balance"]

    # Row 18 = "Hossein" / Person / IMPORTRANGE from a private external sheet.
    # No template user can access that sheet -- remove the row entirely.
    ws.delete_rows(18, 1)
    # Former row 19 ("Ehsan", Person, Closed) is now row 18 after the delete;
    # genericize the real name, keep the row as a demo of the "Person" account type.
    ws.cell(row=18, column=1, value="Friend")

    # Demo balances (replace personal real-money figures with small round numbers).
    demo_balances = {
        3: 160.0,      # Cash
        4: 5000.0,     # WS (Wealthsimple Saving)
        6: 800.0,      # RRSP (Member) -- offset removed, formula below sums demo contributions
        8: 1200.0,     # Regular investment
        9: 600.0,      # TFSA (Managed)
        11: 300.0,     # TFSA
        12: -50.0,     # WV credit
    }
    for row, value in demo_balances.items():
        cell = ws.cell(row=row, column=4)
        if not isinstance(cell.value, str) or not cell.value.startswith("="):
            cell.value = value

    # Unwrap the xlsx-export artifact wrapping GOOGLEFINANCE; works natively once back in Sheets.
    ws.cell(row=5, column=4, value='=round(150*GOOGLEFINANCE("CURRENCY:USDCAD"),2)')

    # Drop the personal opening-balance literal baked into the RRSP (Member) formula.
    ws.cell(row=6, column=4, value='=SUMIF(Transactions!B:B, "RRSP (Member)", Transactions!E:E)')

    # FHSA / CIBCM / TDV / TWM were personal manual-adjustment formulas that net to a
    # hardcoded figure or to zero -- replace with plain demo/zero values.
    ws.cell(row=10, column=4, value=400.0)   # FHSA (Managed)
    ws.cell(row=13, column=4, value=0.0)     # CIBCM
    ws.cell(row=14, column=4, value=0.0)     # TDV
    ws.cell(row=16, column=4, value=0.0)     # TWM


def rebuild_insight(wb):
    ws = wb["Insight"]
    # All per-category formulas reference Transactions!...71 (the row bound that
    # matched whatever Transactions size existed when this sheet was built).
    # Re-bound to the new demo data's last row.
    old_bound = "71"
    new_bound = str(NEW_LAST_ROW)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula = cell.value
                for col_letter in ("A", "D", "E", "F"):
                    formula = formula.replace(f"${col_letter}{old_bound}", f"${col_letter}{new_bound}")
                cell.value = formula


def rebuild_reconciliation(wb):
    ws = wb["Reconciliation"]
    # Unwrap the xlsx-export artifact wrapping FILTER; works natively in Sheets.
    ws.cell(
        row=2, column=2,
        value="=IFERROR(FILTER('Monthly Summary'!N2:N6, "
              "EOMONTH(DATEVALUE('Monthly Summary'!A2:A6&\"-01\"),0)=EOMONTH(TODAY(),0)),0)",
    )
    # Personal investment-interest figure -> demo value.
    ws.cell(row=3, column=2, value=120.0)


def main():
    wb = openpyxl.load_workbook(SOURCE_PATH, data_only=False)
    rebuild_transactions(wb)
    rebuild_account_balance(wb)
    rebuild_insight(wb)
    rebuild_reconciliation(wb)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
