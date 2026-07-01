"""
Merge contacts.csv (Google Contacts export), TELL.xlsx (old manual sheet), and
people.json (resume contacts) into one deduplicated workbook ready to paste
into the Ledger Google Sheet's "Contacts" tab (see README.md Data Model).

Reads all three source files read-only. Writes OUTPUT_PATH only.

Output has 2 sheets:
  - "Contacts"          the 21-column merged, deduplicated contact list.
                         Anything that didn't fit a column slot (4th+ phone,
                         3rd+ email, 2nd+ website/LinkedIn, 3rd+ Telegram,
                         unparsable source values) is folded into that
                         contact's own Note field -- nothing is silently
                         dropped, but nothing needs its own sparse column either.
  - "Possible Duplicates"  near-identical names the automatic name/phone
                         matching below did not merge, flagged for manual review
"""

import csv
import json
import re
import difflib
from collections import defaultdict

import openpyxl

GOOGLE_CSV_PATH = r"C:\Users\AlanKay\Downloads\contacts.csv"
TELL_XLSX_PATH = r"C:\Users\AlanKay\Downloads\TELL.xlsx"
PEOPLE_JSON_PATH = r"C:\Users\AlanKay\Downloads\ainlamyae.github.io\assets\data\people.json"
OUTPUT_PATH = r"C:\Users\AlanKay\Downloads\contacts_merged.xlsx"

HEADERS = [
    "First Name", "Middle Name", "Last Name", "Prefix", "Tags", "Birthday",
    "Phone 1", "Phone 2", "Phone 3", "Email 1", "Email 2",
    "Street Address", "City", "Province/Region", "Postal Code", "Country",
    "Website", "LinkedIn", "Telegram", "Telegram 2", "Note",
]

TITLE_DISPLAY = {"dr": "Dr.", "prof": "Prof.", "mr": "Mr.", "mrs": "Mrs.", "ms": "Ms.", "eng": "Eng."}

PHONE_MIN_DIGITS = 7

# Some contacts have their phone number typed into a free-text Notes field
# instead of a Phone field -- these patterns catch that without also
# matching unrelated digit strings (e.g. bank account numbers in Notes).
PHONE_IN_TEXT_RE = re.compile(
    r'(?:\+|00)\d[\d\s\-]{7,}\d'          # +98 913 168 7313 / 0098 21 8802 8520
    r'|\(\d{3}\)\s?\d{3}[\s.\-]?\d{4}'    # (416) 333-8907
    r'|\d{3}[\s.\-]\d{3}[\s.\-]\d{4}'     # 416-333-8907 / 416.333.8907
)


class Contact:
    def __init__(self):
        self.first = ""
        self.middle = ""
        self.last = ""
        self.prefix = ""
        self.birthday = ""
        self.phones = []
        self.emails = []
        self.street = ""
        self.city = ""
        self.region = ""
        self.postal = ""
        self.country = ""
        self.website = []
        self.linkedin = []
        self.telegram = []
        self.tags = []
        self.notes = []

    @property
    def name(self):
        return f"{self.first} {self.last}".strip()

    def add_phone(self, raw):
        raw = raw.strip()
        d = last10(raw)
        if not d or any(last10(p) == d for p in self.phones):
            return
        self.phones.append(raw)

    def add_email(self, raw):
        raw = raw.strip()
        if '@' not in raw or any(e.lower() == raw.lower() for e in self.emails):
            return
        self.emails.append(raw)

    def add_url(self, raw):
        raw = raw.strip()
        if not raw:
            return
        bucket = getattr(self, route_url(raw))
        if any(u.lower() == raw.lower() for u in bucket):
            return
        bucket.append(raw)

    def add_tag(self, tag):
        tag = (tag or '').strip()
        if tag and tag.lower() not in (t.lower() for t in self.tags):
            self.tags.append(tag)

    def add_note(self, text):
        text = (text or '').strip()
        if text and text not in self.notes:
            self.notes.append(text)


def strip_title(name):
    """Split a leading title token (Dr/Prof/Mr/Mrs/Ms/Eng) off a first-name string."""
    name = (name or '').strip()
    if not name:
        return '', ''
    parts = name.split(' ', 1)
    token = parts[0].rstrip('.').lower()
    if token in TITLE_DISPLAY:
        return TITLE_DISPLAY[token], (parts[1] if len(parts) > 1 else '').strip()
    return '', name


def normalize_key(first, last):
    clean = lambda s: re.sub(r'[^a-z]', '', (s or '').lower())
    return f"{clean(first)}|{clean(last)}"


def digits_only(s):
    return re.sub(r'\D', '', s or '')


def last10(s):
    d = digits_only(s)
    return d[-10:] if len(d) >= PHONE_MIN_DIGITS else ''


def route_url(url):
    low = url.lower()
    if 't.me' in low or 'telegram.me' in low or 'telegram.org' in low:
        return 'telegram'
    if 'linkedin.com' in low:
        return 'linkedin'
    return 'website'


def split_multi(value):
    """Google export packs multiple values into one cell separated by ' ::: '."""
    if not value:
        return []
    return [v.strip() for v in value.split(' ::: ') if v.strip()]


def build_tags(org_name, labels_raw):
    tags = []
    if org_name and org_name.strip():
        tags.append(org_name.strip())
    for lbl in split_multi(labels_raw):
        lbl = lbl.lstrip('* ').strip()
        if lbl and lbl.lower() not in ('mycontacts', 'other'):
            tags.append(lbl)
    seen, out = set(), []
    for t in tags:
        if t.lower() not in seen:
            seen.add(t.lower())
            out.append(t)
    return out


def load_google(path):
    contacts = []
    with open(path, encoding='utf-8-sig', newline='') as f:
        for row in csv.DictReader(f):
            c = Contact()
            raw_first = (row.get('First Name') or '').strip()
            prefix_col = (row.get('Name Prefix') or '').strip()
            if prefix_col:
                c.prefix, c.first = prefix_col, raw_first
            else:
                c.prefix, c.first = strip_title(raw_first)
            c.middle = (row.get('Middle Name') or '').strip()
            c.last = (row.get('Last Name') or '').strip()

            bday = (row.get('Birthday') or '').strip()
            if re.match(r'^\d{4}-\d{2}-\d{2}$', bday):
                c.birthday = bday
            elif bday:
                c.add_note(f"Birthday (unrecognized format): {bday}")

            for col in ('Phone 1 - Value', 'Phone 2 - Value'):
                for val in split_multi(row.get(col)):
                    if last10(val):
                        c.add_phone(val)
                    else:
                        c.add_note(f"Invalid phone: {val}")

            notes = (row.get('Notes') or '').strip()
            for m in PHONE_IN_TEXT_RE.findall(notes):
                c.add_phone(m)
            leftover = PHONE_IN_TEXT_RE.sub('', notes).strip(' \n,')
            if leftover:
                c.add_note(leftover)

            for col in ('E-mail 1 - Value', 'E-mail 2 - Value'):
                for val in split_multi(row.get(col)):
                    if '@' in val:
                        c.add_email(val)
                    else:
                        c.add_note(f"Invalid email: {val}")

            c.street = (row.get('Address 1 - Street') or '').strip()
            c.city = (row.get('Address 1 - City') or '').strip()
            c.region = (row.get('Address 1 - Region') or '').strip()
            c.postal = (row.get('Address 1 - Postal Code') or '').strip()
            c.country = (row.get('Address 1 - Country') or '').strip()
            po_box = (row.get('Address 1 - PO Box') or '').strip()
            if po_box:
                c.street = f"{c.street} (PO Box {po_box})".strip()
            ext = (row.get('Address 1 - Extended Address') or '').strip()
            if ext:
                c.street = f"{c.street}, {ext}".strip(', ')
            addr2 = (row.get('Address 2 - Formatted') or '').strip()
            if addr2:
                c.add_note(f"Address 2: {addr2}")

            for col in ('Website 1 - Value', 'Website 2 - Value'):
                for val in split_multi(row.get(col)):
                    c.add_url(val)

            for t in build_tags(row.get('Organization Name'), row.get('Labels')):
                c.add_tag(t)

            if c.first or c.last:
                contacts.append(c)
    return contacts


def load_tell(path):
    contacts = []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        first_raw = str(row[0]).strip() if row[0] else ''
        last_raw = str(row[1]).strip() if row[1] else ''
        if not first_raw and not last_raw:
            continue
        c = Contact()
        c.prefix, c.first = strip_title(first_raw)
        c.last = last_raw
        for val in row[2:6]:
            if val is None:
                continue
            val = str(val).strip()
            if not val:
                continue
            if len(digits_only(val)) >= PHONE_MIN_DIGITS:
                c.add_phone(val)
            else:
                c.add_note(f"Invalid phone (TELL): {val}")
        contacts.append(c)
    return contacts


def load_people_json(path):
    contacts = []
    with open(path, encoding='utf-8') as f:
        data = json.load(f)
    for p in data:
        c = Contact()
        c.first = (p.get('givenName') or '').strip()
        c.middle = (p.get('middleName') or '').strip()
        c.last = (p.get('familyName') or '').strip()
        prefix = (p.get('prefix') or '').strip()
        if prefix.startswith('http'):
            # Data bug in this source: a couple of entries have a LinkedIn URL
            # sitting in the "prefix" field instead of "url".
            c.add_url(prefix)
        else:
            c.prefix = prefix
        email = (p.get('email') or '').strip()
        if email:
            c.add_email(email)
        phone = (p.get('phone') or '').strip()
        if phone:
            if len(digits_only(phone)) >= PHONE_MIN_DIGITS:
                c.add_phone(phone)
            else:
                c.add_note(f"Invalid phone (Resume): {phone}")
        url = (p.get('url') or '').strip()
        if url:
            c.add_url(url)
        if c.first or c.last:
            contacts.append(c)
    return contacts


def merge_into(target, source):
    for field in ('prefix', 'middle', 'birthday', 'street', 'city', 'region', 'postal', 'country'):
        if not getattr(target, field) and getattr(source, field):
            setattr(target, field, getattr(source, field))
    for p in source.phones:
        target.add_phone(p)
    for e in source.emails:
        target.add_email(e)
    for u in source.website + source.linkedin + source.telegram:
        target.add_url(u)
    for t in source.tags:
        target.add_tag(t)
    for n in source.notes:
        target.add_note(n)


def merge_all(google_contacts, tell_contacts, people_contacts):
    contacts = {}
    for c in google_contacts:
        key = normalize_key(c.first, c.last)
        if key in contacts:
            merge_into(contacts[key], c)
        else:
            contacts[key] = c

    def phone_index():
        idx = defaultdict(list)
        for key, c in contacts.items():
            for p in c.phones:
                d = last10(p)
                if d:
                    idx[d].append(key)
        return idx

    stats = {'tell_matched': 0, 'tell_new': 0, 'people_matched': 0, 'people_new': 0}

    for c in tell_contacts:
        key = normalize_key(c.first, c.last)
        target_key = key if key in contacts else None
        if target_key is None:
            idx = phone_index()
            for p in c.phones:
                d = last10(p)
                if d and idx.get(d):
                    target_key = idx[d][0]
                    break
        if target_key:
            merge_into(contacts[target_key], c)
            stats['tell_matched'] += 1
        else:
            contacts[key] = c
            stats['tell_new'] += 1

    for c in people_contacts:
        key = normalize_key(c.first, c.last)
        if key in contacts:
            merge_into(contacts[key], c)
            stats['people_matched'] += 1
        else:
            contacts[key] = c
            stats['people_new'] += 1

    return contacts, stats


def find_possible_duplicates(contacts, threshold=0.82):
    # Name-string similarity only -- a best-effort safety net, not exhaustive.
    # Cross-source spelling drift on *both* name parts at once (rare) can slip past this.
    items = [(c.name, c) for c in contacts.values() if c.name]
    pairs = []
    for i in range(len(items)):
        n1, c1 = items[i]
        for j in range(i + 1, len(items)):
            n2, c2 = items[j]
            if c1 is c2:
                continue
            ratio = difflib.SequenceMatcher(None, n1.lower(), n2.lower()).ratio()
            if ratio >= threshold:
                pairs.append((n1, n2, round(ratio, 2)))
    pairs.sort(key=lambda x: -x[2])
    return pairs


def finalize_row(contact):
    for extra in contact.phones[3:]:
        contact.add_note(f"Extra phone: {extra}")
    for extra in contact.emails[2:]:
        contact.add_note(f"Extra email: {extra}")
    for extra in contact.website[1:]:
        contact.add_note(f"Extra website: {extra}")
    for extra in contact.linkedin[1:]:
        contact.add_note(f"Extra LinkedIn: {extra}")
    for extra in contact.telegram[2:]:
        contact.add_note(f"Extra Telegram: {extra}")

    g = lambda lst, i: lst[i] if i < len(lst) else ''
    return [
        contact.first, contact.middle, contact.last, contact.prefix,
        ', '.join(contact.tags), contact.birthday,
        g(contact.phones, 0), g(contact.phones, 1), g(contact.phones, 2),
        g(contact.emails, 0), g(contact.emails, 1),
        contact.street, contact.city, contact.region, contact.postal, contact.country,
        g(contact.website, 0), g(contact.linkedin, 0), g(contact.telegram, 0), g(contact.telegram, 1),
        '\n'.join(contact.notes),
    ]


def write_output(contacts, dup_pairs, path):
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Contacts'
    ws.append(HEADERS)
    for c in sorted(contacts.values(), key=lambda c: (c.last.lower(), c.first.lower())):
        ws.append(finalize_row(c))

    ws_dupes = wb.create_sheet('Possible Duplicates')
    ws_dupes.append(['Name A', 'Name B', 'Similarity'])
    for a, b, ratio in dup_pairs:
        ws_dupes.append([a, b, ratio])

    wb.save(path)


def main():
    google_contacts = load_google(GOOGLE_CSV_PATH)
    tell_contacts = load_tell(TELL_XLSX_PATH)
    people_contacts = load_people_json(PEOPLE_JSON_PATH)

    contacts, stats = merge_all(google_contacts, tell_contacts, people_contacts)
    # Possible-duplicate detection must run before finalize_row (called inside
    # write_output) starts mutating each contact's notes with overflow text.
    dup_pairs = find_possible_duplicates(contacts)

    write_output(contacts, dup_pairs, OUTPUT_PATH)

    notes_count = sum(len(c.notes) for c in contacts.values())

    print(f"Google contacts:  {len(google_contacts)}")
    print(f"TELL rows:        {len(tell_contacts)} ({stats['tell_matched']} matched, {stats['tell_new']} new)")
    print(f"Resume entries:   {len(people_contacts)} ({stats['people_matched']} matched, {stats['people_new']} new)")
    print(f"Merged contacts:  {len(contacts)}")
    print(f"Note entries:     {notes_count} (overflow values folded into each contact's Note field)")
    print(f"Possible dupes:   {len(dup_pairs)}")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == '__main__':
    main()
